from openpyxl import Workbook
import os
import time
import re
import requests
import json
import random

headers={'Content-Type':'application/json;charset=UTF-8',
       'Referer':'https://test.swftcoin.com/swft-v3/swft-v3-pc/login.html',
'User-Agent':"Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36"}
#用户登陆加载cookies和存储session
def swft_session():
    swftlogin_url='https://test.swftcoin.com/accountapi/login_web'
    login_payload={"equipmentNo":"QM8BF1tXTnpqMQyZ4b4bn8K9qL2Va6xb",
    "sessionUuid":"","sourceType":"H5","userNo":'1211024726@qq.com',"userPassword":'1qaz2wsx'}
    swftloginSession=requests.session()
    swftlogin_req1=swftloginSession.post(swftlogin_url,data=json.dumps(login_payload),headers=headers).text
    swftlogin_req2=json.loads(swftlogin_req1)
    return swftloginSession

def queryCoinList():
    queryCoinListUrl='https://test.swftcoin.com/api/v1/queryCoinList'
    queryCoinList_payload={}
    session=swft_session()
    res1=session.post(queryCoinListUrl,
                        data=json.dumps(queryCoinList_payload),headers=headers,allow_redirects=False).text
    res2=json.loads(res1)
    global queryCoinList
    queryCoinList=res2
    # print(queryCoinList)
    global num
    num=len(queryCoinList['data'])
    # print(num)

def getBaseInfo():
    baseInfoUrl='https://test.swftcoin.com/api/v1/getBaseInfo'
    baseInfo_payload={'depositCoinCode': upper_dep,
                        'receiveCoinCode': upper_rec
                    }
    session=swft_session()
    res1=session.post(baseInfoUrl,
                        data=json.dumps(baseInfo_payload),headers=headers,allow_redirects=False).text
    res2=json.loads(res1)

    # try:
    #     global depCoininstantRate
    #     depCoininstantRate=res2['data']['instantRate']
    #     for i in range(len(depCoininstantRate)):
    #         print(depCoininstantRate[i])
    #     global depositMax
    #     depositMax=float(res2['data']['depositMax'])
    #     global depositMin
    #     depositMin=float(res2['data']['depositMin'])
    # except:
    #     print(res2)
    #     currentlist.append(res2['resMsg'])
    #
    # try:
    #     global recCoininsantRate
    #     recCoininsantRate=res2['data']['instantRate']['receiveCoinDepth']
    # #print(recCoininsantRate)
    #     for i in range(len(recCoininsantRate)):
    #         print(recCoininsantRate[i])
    # except:
    #     print(res2)
    try:
        global instantRate
        instantRate=float(res2['data']['instantRate'])
        global depositMax
        depositMax=float(res2['data']['depositMax'])
        global depositMin
        depositMin=float(res2['data']['depositMin'])

    except:
        print(res2)

def exchangeAm():
    try:
        global exchangeAmount
        exchangeAmount=random.uniform(depositMin,depositMax)+depositMin
        # print('存币数量:',exchangeAmount)
        # currentlist.append(exchangeAmount)
    except:
        print('原因参见上条')
        # currentlist.append('原因参见上条')

    # global exchangeAmount
    # exchangeAmount=float(input('存入币种兑换值:'))#需要定义
    # print('存入币种兑换值:',exchangeAmount)
    #判断存入币种深度及汇率
    return

# def exchangeRate():
#     try:
#         # if exchangeAmount>depCoininstantRate[len(depCoininstantRate)-1][1]:
        #     print('超出存币范围1')
        #     # curentlist.append('超出存币范围1')
        # else:#存币开始是数量校验
        #     for amount in range(len(depCoininstantRate)):
        #         if exchangeAmount<=depCoininstantRate[amount][1]:
        #             global exchangeRate1
        #             exchangeRate1=depCoininstantRate[amount][0]
        #             global Amount1
        #             Amount1=exchangeRate1*exchangeAmount
        #             print('存币币种汇率:',exchangeRate1)
        #             currentlist.append(exchangeRate1)
        #             break
        #         else:
        #             continue
        #     try:
        #         if Amount1>recCoininsantRate[len(recCoininsantRate)-1][2]:
        #             print('超出存币范围2')
        #             currentlist.append('超出存币范围2')
        #         else:
        #             for amount in range(len(recCoininsantRate)):
        #                 if Amount1<recCoininsantRate[amount][2]:
        #                     global exchangeRate2
        #                     exchangeRate2=recCoininsantRate[amount][0]
        #                     print('目标币种汇率:',exchangeRate2)
        #                     currentlist.append(exchangeRate2)
        #                     break
        #                 else:
        #                     continue
    # rate=exchangeRate1/exchangeRate2
    # currentlist.append(rate)
    # print(rate)
    # return rate
    #         except:
    #             print('无法交易1')
    #             currentlist.append('无法交易1')
    #             return "null"
    # except:
    #     print('无法交易2')
    #     currentlist.append('无法交易2')
    #     return "null"

def document():
    wb=Workbook()
    wb.create_sheet('1',index=0)
    worksheet=wb['1']
    if start==1:
        title_list=['序号','存币','目标币','存币数量','3.7.0-测试汇率','旧系统汇率','误差']
        for i in range(7):
            worksheet.cell(row=1,column=i+1,value=title_list[i])
    else:
        if start==3:
            worksheet.cell(row=start+1,column=k+1,value=currentlist[k])
            wb.save(r'C:C:\Users\Administrator\Desktop\SWFT\汇率测试.xlsx')
        else:
            for k in range(7):
                worksheet.cell(row=start+1,column=k+1,value=currentlist[k])

def realBaseinfo():
    realBaseinforUrl='https://transfer.swft.pro/api/v1/getBaseInfo'
    realBaseinfor_payload={'depositCoinCode': upper_dep,
                        'receiveCoinCode': upper_rec
                    }
    session=swft_session()
    res1=session.post(realBaseinforUrl,
                        data=json.dumps(realBaseinfor_payload),headers=headers,allow_redirects=False).text
    res2=json.loads(res1)
    global realInstantRate
    try:
        realInstantRate=float(res2['data']['instantRate'])
    except:
        realInstantRate=1

if __name__=='__main__':
    wb=Workbook()
    wb.create_sheet('1',index=0)
    worksheet=wb['1']
    currentlist=[]
    start=1
    swft_session()#必须保留
    queryCoinList()
    for i in range(num):
        upper_dep=queryCoinList['data'][i]['coinCode']
        # upper_dep='SWFTC'
        for k in range(num):
            # print('存币:',upper_dep)
            # for m in range(1):
            upper_rec=queryCoinList['data'][k]['coinCode']
            # upper_rec='DAI'
            if upper_dep==upper_rec:
                continue
            else:
            # currentlist.append(start)
            # currentlist.append(upper_dep)
            # currentlist.append(upper_rec)
            # print('目标币:',upper_rec)
                getBaseInfo()
                realBaseinfo()
                exchangeAm()
                # exchangeRate()
                # try:
                diff=float((instantRate-realInstantRate)/realInstantRate)
                print(start,upper_dep,upper_rec,'存币数量:',exchangeAmount,'测试汇率:',instantRate,'生产汇率:',realInstantRate,'误差:',diff)
                currentlist=[start,upper_dep,upper_rec,exchangeAmount,instantRate,realInstantRate,diff]
                if start==1:
                    title_list=['序号','存币','目标币','存币数量','3.7.0-预接受测试汇率','旧系统预接受汇率','误差']
                    for i in range(7):
                        worksheet.cell(row=1,column=i+1,value=title_list[i])
                        worksheet.cell(row=start+1,column=i+1,value=currentlist[i])
                else:
                    for m in range(7):
                        worksheet.cell(row=start+1,column=m+1,value=currentlist[m])
                # except:
                #     print(start,upper_dep,upper_rec,'存币数量:',exchangeAmount,'测试汇率:',rate,'生产汇率:',realInstantRate)# print(currentlist)
                currentlist=[]
                start=start+1
    wb.save(r'C:\Users\Administrator\Desktop\SWFT\汇率测试.xlsx')
